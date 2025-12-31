"""
HUL Processor - Streamlit compatible version
This is a wrapper that adapts the HUL bulk upload script for Streamlit.
"""
import sys
import os
from pathlib import Path

# Import utilities from the original script
scripts_dir = Path(__file__).parent.parent / "scripts"
sys.path.insert(0, str(scripts_dir))

# Import the original processing function
# We'll need to modify it slightly to work with Streamlit
import pandas as pd
import numpy as np
from fuzzywuzzy import process, fuzz
from collections import Counter
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import streamlit as st

# Copy utility functions
def normalize_name(name):
    if pd.isna(name):
        return ""
    cleaned = " ".join(str(name).strip().split())
    return cleaned

def fuzzy_match_name(name, choices, min_score=0):
    normalized_input = normalize_name(name)
    if not normalized_input:
        return "", 0
    normalized_choices = {normalize_name(c): c for c in choices if normalize_name(c)}
    normalized_choice_list = list(normalized_choices.keys())
    result = process.extractOne(normalized_input, normalized_choice_list, scorer=fuzz.token_sort_ratio)
    if result:
        match_normalized, score = result
        match = normalized_choices.get(match_normalized, match_normalized)
        if score >= min_score:
            return match, score
        else:
            return "", 0
    return "", 0

def exact_match_name(name, choices):
    normalized_input = normalize_name(name).lower()
    if not normalized_input:
        return "", 0
    for choice in choices:
        normalized_choice = normalize_name(choice).lower()
        if normalized_input == normalized_choice:
            return choice, 100
    return "", 0

def safe_read_excel(file_path, **kwargs):
    ext = os.path.splitext(file_path)[-1].lower()
    if ext == ".xls":
        return pd.read_excel(file_path, engine="xlrd", **kwargs)
    elif ext == ".xlsx":
        return pd.read_excel(file_path, engine="openpyxl", **kwargs)
    else:
        raise ValueError(f"Unsupported file type: {ext}")

def process_britannia_sales(input_file, reference_file, output_file, column_mapping=None, selected_state=None):
    """
    Process HUL sales orders - Streamlit compatible version.
    
    Args:
        input_file: Path to input Excel file
        reference_file: Path to reference Excel file
        output_file: Path to output Excel file
        column_mapping: Dict of column mappings (if None, will use defaults)
        selected_state: State to filter merchants (if None, will use first state)
    """
    PARTIAL_MATCH_MIN_SCORE = 70
    
    # Load sales file
    df = safe_read_excel(input_file)
    
    # Column mapping
    if column_mapping is None:
        expected = {
            "order_id": "Bill Number",
            "dms_invoice": "Bill Number",
            "order_date": "Bill Date",
            "product_name": "Product Description",
            "merchant_name": "Party",
            "quantity": "Units",
            "selling_price": "Net Sales",
            "low_price_reason": "Low Price Reason",
            "buyer_branch_id": "Branch ID",
            "warehouse_name": "Warehouse Name"
        }
        
        # Auto-detect columns
        mapping = {}
        optional_columns = ["low_price_reason", "buyer_branch_id", "warehouse_name"]
        
        for key, default in expected.items():
            if key in optional_columns:
                if default in df.columns:
                    mapping[key] = default
                else:
                    mapping[key] = None
            else:
                if default in df.columns:
                    mapping[key] = default
                else:
                    # Try to find similar column
                    found = False
                    for col in df.columns:
                        if key.lower().replace("_", " ") in str(col).lower() or str(col).lower() in key.lower().replace("_", " "):
                            mapping[key] = col
                            found = True
                            break
                    if not found:
                        raise ValueError(f"Could not find column for {key}. Please provide column_mapping.")
    else:
        mapping = column_mapping
    
    # Detect due_date and GST columns for Britannia
    due_date_col = None
    cgst_col = None
    sgst_col = None
    igst_col = None
    for col in df.columns:
        col_lower = str(col).strip().lower()
        if "due" in col_lower:
            due_date_col = col
        if "cgst" in col_lower and "%" in str(col):
            cgst_col = col
        elif ("sgst" in col_lower or "ugst" in col_lower) and "%" in str(col):
            sgst_col = col
        elif "igst" in col_lower and "%" in str(col):
            igst_col = col
    
    # Parse dates with dayfirst=True for Britannia
    def parse_date(date_series):
        parsed = pd.to_datetime(date_series, dayfirst=True, errors="coerce")
        if parsed.isna().any():
            mask = parsed.isna()
            parsed_alt = pd.to_datetime(date_series[mask], errors="coerce")
            parsed = parsed.fillna(parsed_alt)
        return parsed
    
    # Build sale_order_df
    sale_order_df = pd.DataFrame({
        "order_id": df[mapping["order_id"]],
        "dms_invoice": df[mapping["dms_invoice"]],
        "order_date": parse_date(df[mapping["order_date"]]).dt.strftime("%d/%m/%Y"),
        "product_name": df[mapping["product_name"]].astype(str).str.strip(),
        "merchant_name": df[mapping["merchant_name"]].astype(str).str.strip(),
        "quantity": pd.to_numeric(df[mapping["quantity"]], errors="coerce"),
        "net_sales": pd.to_numeric(df[mapping["selling_price"]], errors="coerce"),
        "low_price_reason": df[mapping.get("low_price_reason")].astype(str) 
                        if mapping.get("low_price_reason") and mapping.get("low_price_reason") in df.columns 
                        else pd.Series(["low price"] * len(df)),
        "buyer_branch_id": df[mapping.get("buyer_branch_id")].astype(str) 
                        if mapping.get("buyer_branch_id") and mapping.get("buyer_branch_id") in df.columns 
                        else pd.Series([""] * len(df)),
        "warehouse_name": df[mapping.get("warehouse_name")].astype(str) 
                        if mapping.get("warehouse_name") and mapping.get("warehouse_name") in df.columns 
                        else pd.Series([""] * len(df))
    })
    
    if due_date_col:
        sale_order_df["due_date"] = parse_date(df[due_date_col]).dt.strftime("%d/%m/%Y")
    else:
        sale_order_df["due_date"] = ""
    
    # Add GST columns for Britannia
    if cgst_col:
        sale_order_df["CGST %"] = pd.to_numeric(df[cgst_col], errors="coerce")
    else:
        sale_order_df["CGST %"] = ""
    if sgst_col:
        sale_order_df["SGST / UGST %"] = pd.to_numeric(df[sgst_col], errors="coerce")
    else:
        sale_order_df["SGST / UGST %"] = ""
    if igst_col:
        sale_order_df["IGST %"] = pd.to_numeric(df[igst_col], errors="coerce")
    else:
        sale_order_df["IGST %"] = ""
    
    # Drop incomplete rows
    before = len(sale_order_df)
    sale_order_df = sale_order_df.dropna(subset=["order_id", "order_date", "product_name", "merchant_name"])
    if st:
        st.info(f"Dropped {before - len(sale_order_df)} incomplete rows")
    
    # Calculate selling_price
    sale_order_df["selling_price"] = sale_order_df.apply(
        lambda row: row["net_sales"] / row["quantity"] if row["quantity"] != 0 and not pd.isna(row["quantity"]) and not pd.isna(row["net_sales"]) else 0,
        axis=1
    )
    
    # Load reference sheets
    product_df = safe_read_excel(reference_file, sheet_name="Product Details", dtype={"batch_id": str})
    merchant_df = safe_read_excel(reference_file, sheet_name="merchant_data")
    
    # Clean batch_id
    if "batch_id" in product_df.columns:
        product_df["batch_id"] = product_df["batch_id"].apply(lambda x: str(x).split(".")[0] if pd.notna(x) else "")
    
    # State selection
    unique_states = merchant_df["shop_state"].dropna().unique().tolist()
    if not unique_states:
        raise RuntimeError("No states found in merchant_data.shop_state")
    
    if selected_state is None:
        selected_state = unique_states[0]
    
    merchant_df = merchant_df[merchant_df["shop_state"] == selected_state].copy()
    
    # Product matching - simplified: auto-accept all matches, categorize later
    product_names = product_df["product_name"].dropna().astype(str).str.strip().tolist()
    matched_products, product_scores, user_confirmed_flags = [], [], []
    
    progress_bar = st.progress(0) if st else None
    status_text = st.empty() if st else None
    
    for idx, prod in enumerate(sale_order_df["product_name"]):
        match, score = fuzzy_match_name(prod, product_names, min_score=0)
        user_confirmed = False
        
        # Auto-accept 100% matches, mark 70-99% as needing review
        if score >= 100:
            user_confirmed = True
        elif 70 <= score < 100:
            # For web app, we'll mark these as partial matches
            user_confirmed = False
        
        matched_products.append(match)
        product_scores.append(score)
        user_confirmed_flags.append(user_confirmed)
        
        if progress_bar and status_text:
            progress = (idx + 1) / len(sale_order_df)
            progress_bar.progress(progress)
            status_text.text(f"Matching products: {idx + 1}/{len(sale_order_df)}")
    
    if progress_bar:
        progress_bar.empty()
    if status_text:
        status_text.empty()
    
    sale_order_df["matched_product_name"] = matched_products
    sale_order_df["product_match_score"] = product_scores
    sale_order_df["user_confirmed_match"] = user_confirmed_flags
    
    # Build batch inventory
    batch_inventory = {}
    for _, r in product_df.iterrows():
        pname = str(r["product_name"]).strip()
        batch_inventory.setdefault(pname, []).append({
            "batch_id": str(r.get("batch_id", "")),
            "available_stock": float(r.get("available_stock", np.inf)) if pd.notna(r.get("available_stock", np.nan)) else np.inf,
            "product_id": str(r.get("product_id", "")),
        })
    
    for pname in batch_inventory:
        batch_inventory[pname] = sorted(batch_inventory[pname], key=lambda x: -x["available_stock"])
    
    # Batch allocation - simplified: allocate from first available batch
    sale_order_df = sale_order_df.sort_values(by="order_date").reset_index(drop=True)
    allocated_rows = []
    
    progress_bar2 = st.progress(0) if st else None
    status_text2 = st.empty() if st else None
    
    for idx, row in enumerate(sale_order_df.itertuples()):
        pname = row.matched_product_name
        qty = row.quantity if not pd.isna(row.quantity) else 0
        
        base_row = {
            "order_id": row.order_id,
            "dms_invoice": row.dms_invoice,
            "order_date": row.order_date,
            "product_name": row.product_name,
            "merchant_name": row.merchant_name,
            "net_sales": row.net_sales,
            "low_price_reason": row.low_price_reason,
            "buyer_branch_id": row.buyer_branch_id,
            "warehouse_name": row.warehouse_name,
            "due_date": row.due_date,
            "matched_product_name": row.matched_product_name,
            "product_match_score": row.product_match_score,
            "user_confirmed_match": row.user_confirmed_match,
            "CGST %": sale_order_df.loc[row.Index, "CGST %"] if "CGST %" in sale_order_df.columns else "",
            "SGST / UGST %": sale_order_df.loc[row.Index, "SGST / UGST %"] if "SGST / UGST %" in sale_order_df.columns else "",
            "IGST %": sale_order_df.loc[row.Index, "IGST %"] if "IGST %" in sale_order_df.columns else "",
        }
        
        if not pname or pname not in batch_inventory:
            err = "Product not found in reference"
            allocated_rows.append({
                **base_row,
                "quantity": qty,
                "selling_price": row.net_sales / qty if qty != 0 else 0,
                "batch_id": "",
                "product_id": "",
                "product_error": err
            })
        elif qty < 0:
            # Returns
            batch = batch_inventory[pname][0]
            batch["available_stock"] += abs(qty)
            allocated_rows.append({
                **base_row,
                "quantity": qty,
                "selling_price": row.net_sales / qty if qty != 0 else 0,
                "batch_id": batch["batch_id"],
                "product_id": batch["product_id"],
                "product_error": ""
            })
        elif qty > 0:
            # Allocate from available batches
            remaining_qty = qty
            batch_allocations = []
            
            for batch in batch_inventory[pname]:
                if remaining_qty <= 0:
                    break
                if batch["available_stock"] > 0:
                    qty_from_batch = min(batch["available_stock"], remaining_qty)
                    batch["available_stock"] -= qty_from_batch
                    remaining_qty -= qty_from_batch
                    batch_allocations.append((batch["batch_id"], batch["product_id"], qty_from_batch))
            
            if remaining_qty <= 0:
                for batch_id, product_id, allocated_qty in batch_allocations:
                    portion_selling_price = (row.net_sales * allocated_qty / qty) if qty != 0 else 0
                    allocated_rows.append({
                        **base_row,
                        "quantity": allocated_qty,
                        "selling_price": portion_selling_price / allocated_qty if allocated_qty != 0 else 0,
                        "batch_id": batch_id,
                        "product_id": product_id,
                        "product_error": ""
                    })
            else:
                fulfilled_qty = qty - remaining_qty
                err = f"Insufficient stock: need {int(qty)}, only {int(fulfilled_qty)} available"
                allocated_rows.append({
                    **base_row,
                    "quantity": qty,
                    "selling_price": row.net_sales / qty if qty != 0 else 0,
                    "batch_id": "",
                    "product_id": "",
                    "product_error": err
                })
        else:
            allocated_rows.append({
                **base_row,
                "quantity": qty,
                "selling_price": 0,
                "batch_id": "",
                "product_id": "",
                "product_error": "Zero quantity"
            })
        
        if progress_bar2 and status_text2:
            progress = (idx + 1) / len(sale_order_df)
            progress_bar2.progress(progress)
            status_text2.text(f"Allocating batches: {idx + 1}/{len(sale_order_df)}")
    
    if progress_bar2:
        progress_bar2.empty()
    if status_text2:
        status_text2.empty()
    
    sale_order_df = pd.DataFrame(allocated_rows)
    
    # Merchant matching
    shop_names = merchant_df["shop_name"].dropna().astype(str).str.strip().tolist()
    merchant_names = merchant_df["merchant_name"].dropna().astype(str).str.strip().tolist()
    
    shop_mobile_map = dict(zip(merchant_df["shop_name"], merchant_df.get("merchant_mobile_number", pd.Series([""]*len(merchant_df)))))
    shop_state_map = dict(zip(merchant_df["shop_name"], merchant_df.get("shop_state", pd.Series([""]*len(merchant_df)))))
    merchant_mobile_map = dict(zip(merchant_df["merchant_name"], merchant_df.get("merchant_mobile_number", pd.Series([""]*len(merchant_df)))))
    merchant_state_map = dict(zip(merchant_df["merchant_name"], merchant_df.get("shop_state", pd.Series([""]*len(merchant_df)))))
    
    matched_merchants, buyer_mobiles, shop_states, merchant_scores, find_status = [], [], [], [], []
    
    for merchant in sale_order_df["merchant_name"]:
        match, score = exact_match_name(merchant, shop_names)
        if score == 100:
            matched_merchants.append(match)
            buyer_mobiles.append(shop_mobile_map.get(match, ""))
            shop_states.append(shop_state_map.get(match, ""))
            merchant_scores.append(score)
            find_status.append("shop_name")
            continue
        
        match2, score2 = exact_match_name(merchant, merchant_names)
        if score2 == 100:
            matched_merchants.append(match2)
            buyer_mobiles.append(merchant_mobile_map.get(match2, ""))
            shop_states.append(merchant_state_map.get(match2, ""))
            merchant_scores.append(score2)
            find_status.append("merchant_name")
        else:
            matched_merchants.append("")
            buyer_mobiles.append("")
            shop_states.append("")
            merchant_scores.append(0)
            find_status.append("not_found")
    
    sale_order_df["matched_shop_name"] = matched_merchants
    sale_order_df["buyer_mobile"] = buyer_mobiles
    sale_order_df["shop_state"] = shop_states
    sale_order_df["merchant_match_score"] = merchant_scores
    sale_order_df["merchant_find_status"] = find_status
    sale_order_df["merchant_error"] = [
        "Merchant not matched" if s == "not_found" else "" for s in sale_order_df["merchant_find_status"]
    ]
    
    # Combine errors
    sale_order_df["error_message"] = sale_order_df.apply(
        lambda r: ", ".join(filter(None, [r.get("product_error", ""), r.get("merchant_error", "")])),
        axis=1
    )
    
    # Separate negative quantities
    negative_qty_df = sale_order_df[sale_order_df["quantity"] < 0].copy()
    sale_order_df = sale_order_df[sale_order_df["quantity"] >= 0].copy()
    
    if not negative_qty_df.empty:
        negative_qty_df["abs_quantity"] = negative_qty_df["quantity"].abs()
        negative_qty_df["abs_net_sales"] = negative_qty_df["net_sales"].abs()
        negative_qty_df["price"] = negative_qty_df.apply(
            lambda row: row["abs_net_sales"] / row["abs_quantity"] 
            if row["abs_quantity"] != 0 and not pd.isna(row["abs_quantity"]) and not pd.isna(row["abs_net_sales"]) 
            else 0,
            axis=1
        )
        negative_qty_df["return_amount"] = negative_qty_df["price"]
        
        reason_col = negative_qty_df["low_price_reason"] if "low_price_reason" in negative_qty_df.columns else pd.Series(["low price"] * len(negative_qty_df))
        sales_return_error_message = negative_qty_df.apply(
            lambda r: ", ".join(filter(None, [r.get("product_error", ""), r.get("merchant_error", "")])),
            axis=1
        )
        product_name_col = negative_qty_df.get("matched_product_name", negative_qty_df.get("product_name", pd.Series([""] * len(negative_qty_df))))
        
        sales_return_df = pd.DataFrame({
            "order_id": negative_qty_df["order_id"],
            "product_id": negative_qty_df["product_id"],
            "batch_id": negative_qty_df["batch_id"],
            "product_name": product_name_col.astype(str),
            "price": negative_qty_df["price"],
            "return_qty": negative_qty_df["abs_quantity"],
            "return_amount": negative_qty_df["return_amount"],
            "reason": reason_col.astype(str),
            "error_message": sales_return_error_message.astype(str),
            "sales_return_date": negative_qty_df["order_date"],
            "note": "",
            "remark": ""
        })
    else:
        sales_return_df = pd.DataFrame()
    
    # Categorize orders
    def get_match_category(row):
        product_score = row.get("product_match_score", 0) or 0
        merchant_score = row.get("merchant_match_score", 0) or 0
        user_confirmed = row.get("user_confirmed_match", False)
        has_critical_error = bool(row.get("product_error") and "not found" in str(row.get("product_error", "")).lower()) or bool(row.get("merchant_error"))
        
        if ((product_score == 100 or user_confirmed) and merchant_score == 100 and not has_critical_error):
            return "valid"
        if (PARTIAL_MATCH_MIN_SCORE <= product_score < 100 and merchant_score == 100 and not has_critical_error and not user_confirmed):
            return "partial"
        return "error"
    
    sale_order_df["match_category"] = sale_order_df.apply(get_match_category, axis=1)
    
    rows_with_errors = sale_order_df[sale_order_df["error_message"].astype(str).str.strip() != ""]
    error_order_ids = rows_with_errors["order_id"].unique()
    
    rows_with_partial = sale_order_df[sale_order_df["match_category"] == "partial"]
    partial_order_ids = rows_with_partial["order_id"].unique()
    
    error_df = sale_order_df[sale_order_df["order_id"].isin(error_order_ids)].copy()
    orders_without_errors = sale_order_df[~sale_order_df["order_id"].isin(error_order_ids)].copy()
    partial_df = orders_without_errors[orders_without_errors["order_id"].isin(partial_order_ids)].copy()
    valid_df = orders_without_errors[~orders_without_errors["order_id"].isin(partial_order_ids)].copy()
    
    # Create error messages
    def mk_err_msg(r):
        parts = []
        if r.get("product_error"):
            parts.append(r["product_error"])
        if r.get("merchant_error"):
            parts.append(r["merchant_error"])
        product_score = r.get("product_match_score", 0) or 0
        merchant_score = r.get("merchant_match_score", 0) or 0
        user_confirmed = r.get("user_confirmed_match", False)
        if product_score < 100 and not user_confirmed:
            parts.append(f"Product match {int(product_score)}%")
        if merchant_score < 100:
            parts.append(f"Merchant match {int(merchant_score)}%")
        return ", ".join(parts) if parts else ""
    
    def mk_partial_msg(r):
        parts = []
        product_score = r.get("product_match_score", 0) or 0
        merchant_score = r.get("merchant_match_score", 0) or 0
        user_confirmed = r.get("user_confirmed_match", False)
        if user_confirmed:
            parts.append(f"User confirmed match ({int(product_score)}%)")
        elif product_score < 100:
            parts.append(f"Product match {int(product_score)}%")
        if merchant_score < 100:
            parts.append(f"Merchant match {int(merchant_score)}%")
        if r.get("product_error"):
            parts.append(r["product_error"])
        return ", ".join(parts) if parts else "Partial match"
    
    if not error_df.empty:
        error_df["error_message"] = error_df.apply(mk_err_msg, axis=1)
    
    if not partial_df.empty:
        partial_df["partial_match_reason"] = partial_df.apply(mk_partial_msg, axis=1)
    
    # Reorder columns
    top_cols = ["order_id","order_date","warehouse_name","product_id","batch_id",
                "buyer_mobile","buyer_branch_id","quantity","selling_price","due_date",
                "dms_invoice","low_price_reason","CGST %","SGST / UGST %","IGST %"]
    
    if not valid_df.empty:
        rest_cols = [c for c in valid_df.columns if c not in top_cols]
        valid_df = valid_df[top_cols + rest_cols]
    
    if not partial_df.empty:
        rest_cols_partial = [c for c in partial_df.columns if c not in top_cols]
        partial_df = partial_df[top_cols + rest_cols_partial]
    
    # Save to Excel
    with pd.ExcelWriter(output_file, engine="openpyxl", mode="w") as writer:
        if not valid_df.empty:
            valid_df.to_excel(writer, sheet_name="Sale Order Demo", index=False)
        if not partial_df.empty:
            partial_df.to_excel(writer, sheet_name="Partially Matched", index=False)
        if not error_df.empty:
            error_df.to_excel(writer, sheet_name="Error Rows", index=False)
        if not sales_return_df.empty:
            sales_return_df.to_excel(writer, sheet_name="Sales Return Sheet", index=False)
        product_df.to_excel(writer, sheet_name="Product Details", index=False)
        merchant_df.to_excel(writer, sheet_name="merchant_data", index=False)
    
    # Apply color fills
    try:
        wb = load_workbook(output_file)
        if "Error Rows" in wb.sheetnames:
            ws = wb["Error Rows"]
            red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            for r in range(2, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = red
        if "Partially Matched" in wb.sheetnames:
            ws = wb["Partially Matched"]
            yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            for r in range(2, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = yellow
        wb.save(output_file)
    except Exception as e:
        if st:
            st.warning(f"Could not apply coloring to Excel file: {e}")
    
    return valid_df, partial_df, error_df, sales_return_df

