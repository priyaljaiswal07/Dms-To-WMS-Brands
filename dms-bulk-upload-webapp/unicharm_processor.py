"""
Unicharm Processor - Streamlit compatible version
Based on the original unicharm bulk upload script
"""
import pandas as pd
import numpy as np
from fuzzywuzzy import process, fuzz
import os
from collections import Counter
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import timedelta
import streamlit as st

# Import utility functions from hul_processor
from hul_processor import (
    normalize_name, fuzzy_match_name, exact_match_name, safe_read_excel
)

def process_unicharm_sales(input_file, reference_file, output_file, column_mapping=None, selected_state=None, 
                          warehouse_name="", low_price_reason="low price", buyer_branch_id=""):
    """
    Process Unicharm sales orders - Streamlit compatible version.
    """
    PARTIAL_MATCH_MIN_SCORE = 70
    
    # Load sales file with multi-header
    df = safe_read_excel(input_file, header=[6,7,8])
    
    # Handle MultiIndex columns
    def combine_multiindex_column(col):
        if isinstance(col, tuple):
            values = [str(x).strip() for x in col if str(x).strip().lower() not in ['nan', 'none', '']]
            seen = set()
            unique_values = []
            for v in values:
                if v not in seen:
                    seen.add(v)
                    unique_values.append(v)
            return " ".join(unique_values).strip()
        return str(col).strip()
    
    df.columns = [combine_multiindex_column(col) for col in df.columns.values]
    
    # Column mapping - Unicharm specific
    if column_mapping is None:
        expected = {
            "order_id": ["Invoice Number", "Invoice"],
            "order_date": ["Invoice Date", "Invoice.1"],
            "quantity": ["Total Quantity", "Total"],
            "selling_price": ["Product Level Net Amount", "Taxable", "Net Amount"],
            "dms_invoice": ["Invoice Number", "Invoice"],
            "product_name": ["Product Name"],
            "merchant_name": ["Retailer Name", "Retailer", "Retailer.3"]
        }
        
        mapping = {}
        for key, options in expected.items():
            found = None
            for col in options:
                if col in df.columns:
                    found = col
                    break
            if not found:
                # Use first column as fallback
                found = df.columns[0] if len(df.columns) > 0 else None
            mapping[key] = found
    else:
        mapping = column_mapping
    
    # Detect GST columns
    output_cgst_col = None
    output_sgst_col = None
    for col in df.columns:
        col_str = str(col).strip()
        if "[OutputCGST Rate]" in col_str or "OutputCGST Rate" in col_str or ("outputcgst" in col_str.lower() and "rate" in col_str.lower()):
            output_cgst_col = col
        elif "[OutputSGST Rate]" in col_str or "OutputSGST Rate" in col_str or ("outputsgst" in col_str.lower() and "rate" in col_str.lower()):
            output_sgst_col = col
    
    # Parse dates
    def parse_date(date_series):
        parsed = pd.to_datetime(date_series, dayfirst=True, errors="coerce")
        if parsed.isna().any():
            mask = parsed.isna()
            parsed_alt = pd.to_datetime(date_series[mask], errors="coerce")
            parsed = parsed.fillna(parsed_alt)
        return parsed
    
    # Build sale_order_df
    order_date_parsed = parse_date(df[mapping["order_date"]])
    sale_order_df = pd.DataFrame({
        "order_id": df[mapping["order_id"]],
        "dms_invoice": df[mapping["dms_invoice"]],
        "order_date": order_date_parsed.dt.strftime("%d/%m/%Y"),
        "product_name": df[mapping["product_name"]].astype(str).str.strip(),
        "merchant_name": df[mapping["merchant_name"]].astype(str).str.strip(),
        "quantity": pd.to_numeric(df[mapping["quantity"]], errors="coerce"),
        "net_sales": pd.to_numeric(df[mapping["selling_price"]], errors="coerce"),
        "low_price_reason": low_price_reason if low_price_reason else "low price",
        "buyer_branch_id": buyer_branch_id,
        "warehouse_name": warehouse_name
    })
    
    # Calculate due_date (10 days from order_date)
    sale_order_df["due_date"] = (order_date_parsed + timedelta(days=10)).dt.strftime("%d/%m/%Y")
    
    # Add GST columns
    if output_cgst_col:
        sale_order_df["[OutputCGST Rate]"] = pd.to_numeric(df[output_cgst_col], errors="coerce")
    else:
        sale_order_df["[OutputCGST Rate]"] = ""
    if output_sgst_col:
        sale_order_df["[OutputSGST Rate]"] = pd.to_numeric(df[output_sgst_col], errors="coerce")
    else:
        sale_order_df["[OutputSGST Rate]"] = ""
    
    # Drop incomplete rows
    before = len(sale_order_df)
    sale_order_df = sale_order_df.dropna(subset=["order_id", "order_date", "product_name", "merchant_name"])
    if st:
        st.info(f"Dropped {before - len(sale_order_df)} incomplete rows")
    
    # Calculate selling_price
    sale_order_df["selling_price"] = sale_order_df.apply(
        lambda row: row["net_sales"] / row["quantity"] if row["quantity"] != 0 and not pd.isna(row["quantity"]) and not pd.isna(row["net_sales"]) else 0,
        axis=1
    )
    
    # Load reference sheets
    product_df = safe_read_excel(reference_file, sheet_name="Product Details", dtype={"batch_id": str})
    merchant_df = safe_read_excel(reference_file, sheet_name="merchant_data")
    
    # Clean batch_id
    if "batch_id" in product_df.columns:
        product_df["batch_id"] = product_df["batch_id"].apply(lambda x: str(x).split(".")[0] if pd.notna(x) else "")
    
    # State selection
    unique_states = merchant_df["shop_state"].dropna().unique().tolist()
    if not unique_states:
        raise RuntimeError("No states found in merchant_data.shop_state")
    
    if selected_state is None:
        selected_state = unique_states[0]
    
    merchant_df = merchant_df[merchant_df["shop_state"] == selected_state].copy()
    
    # Product matching - simplified
    product_names = product_df["product_name"].dropna().astype(str).str.strip().tolist()
    matched_products, product_scores, user_confirmed_flags = [], [], []
    
    progress_bar = st.progress(0) if st else None
    status_text = st.empty() if st else None
    
    for idx, prod in enumerate(sale_order_df["product_name"]):
        match, score = fuzzy_match_name(prod, product_names, min_score=0)
        user_confirmed = (score >= 100)
        
        matched_products.append(match)
        product_scores.append(score)
        user_confirmed_flags.append(user_confirmed)
        
        if progress_bar and status_text:
            progress = (idx + 1) / len(sale_order_df)
            progress_bar.progress(progress)
            status_text.text(f"Matching products: {idx + 1}/{len(sale_order_df)}")
    
    if progress_bar:
        progress_bar.empty()
    if status_text:
        status_text.empty()
    
    sale_order_df["matched_product_name"] = matched_products
    sale_order_df["product_match_score"] = product_scores
    sale_order_df["user_confirmed_match"] = user_confirmed_flags
    
    # Build batch inventory
    batch_inventory = {}
    for _, r in product_df.iterrows():
        pname = str(r["product_name"]).strip()
        batch_inventory.setdefault(pname, []).append({
            "batch_id": str(r.get("batch_id", "")),
            "available_stock": float(r.get("available_stock", np.inf)) if pd.notna(r.get("available_stock", np.nan)) else np.inf,
            "product_id": str(r.get("product_id", "")),
        })
    
    for pname in batch_inventory:
        batch_inventory[pname] = sorted(batch_inventory[pname], key=lambda x: -x["available_stock"])
    
    # Batch allocation - simplified
    sale_order_df = sale_order_df.sort_values(by="order_date").reset_index(drop=True)
    allocated_rows = []
    
    progress_bar2 = st.progress(0) if st else None
    status_text2 = st.empty() if st else None
    
    for idx, row in enumerate(sale_order_df.itertuples()):
        pname = row.matched_product_name
        qty = row.quantity if not pd.isna(row.quantity) else 0
        
        base_row = {
            "order_id": row.order_id,
            "dms_invoice": row.dms_invoice,
            "order_date": row.order_date,
            "product_name": row.product_name,
            "merchant_name": row.merchant_name,
            "net_sales": row.net_sales,
            "low_price_reason": row.low_price_reason,
            "buyer_branch_id": row.buyer_branch_id,
            "warehouse_name": row.warehouse_name,
            "due_date": row.due_date,
            "matched_product_name": row.matched_product_name,
            "product_match_score": row.product_match_score,
            "user_confirmed_match": row.user_confirmed_match,
            "[OutputCGST Rate]": sale_order_df.loc[row.Index, "[OutputCGST Rate]"] if "[OutputCGST Rate]" in sale_order_df.columns else "",
            "[OutputSGST Rate]": sale_order_df.loc[row.Index, "[OutputSGST Rate]"] if "[OutputSGST Rate]" in sale_order_df.columns else "",
        }
        
        if not pname or pname not in batch_inventory:
            err = "Product not found in reference"
            allocated_rows.append({
                **base_row,
                "quantity": qty,
                "selling_price": row.net_sales / qty if qty != 0 else 0,
                "batch_id": "",
                "product_id": "",
                "product_error": err
            })
        elif qty < 0:
            batch = batch_inventory[pname][0]
            batch["available_stock"] += abs(qty)
            allocated_rows.append({
                **base_row,
                "quantity": qty,
                "selling_price": row.net_sales / qty if qty != 0 else 0,
                "batch_id": batch["batch_id"],
                "product_id": batch["product_id"],
                "product_error": ""
            })
        elif qty > 0:
            remaining_qty = qty
            batch_allocations = []
            
            for batch in batch_inventory[pname]:
                if remaining_qty <= 0:
                    break
                if batch["available_stock"] > 0:
                    qty_from_batch = min(batch["available_stock"], remaining_qty)
                    batch["available_stock"] -= qty_from_batch
                    remaining_qty -= qty_from_batch
                    batch_allocations.append((batch["batch_id"], batch["product_id"], qty_from_batch))
            
            if remaining_qty <= 0:
                for batch_id, product_id, allocated_qty in batch_allocations:
                    portion_selling_price = (row.net_sales * allocated_qty / qty) if qty != 0 else 0
                    allocated_rows.append({
                        **base_row,
                        "quantity": allocated_qty,
                        "selling_price": portion_selling_price / allocated_qty if allocated_qty != 0 else 0,
                        "batch_id": batch_id,
                        "product_id": product_id,
                        "product_error": ""
                    })
            else:
                fulfilled_qty = qty - remaining_qty
                err = f"Insufficient stock: need {int(qty)}, only {int(fulfilled_qty)} available"
                allocated_rows.append({
                    **base_row,
                    "quantity": qty,
                    "selling_price": row.net_sales / qty if qty != 0 else 0,
                    "batch_id": "",
                    "product_id": "",
                    "product_error": err
                })
        else:
            allocated_rows.append({
                **base_row,
                "quantity": qty,
                "selling_price": 0,
                "batch_id": "",
                "product_id": "",
                "product_error": "Zero quantity"
            })
        
        if progress_bar2 and status_text2:
            progress = (idx + 1) / len(sale_order_df)
            progress_bar2.progress(progress)
            status_text2.text(f"Allocating batches: {idx + 1}/{len(sale_order_df)}")
    
    if progress_bar2:
        progress_bar2.empty()
    if status_text2:
        status_text2.empty()
    
    sale_order_df = pd.DataFrame(allocated_rows)
    
    # Merchant matching
    shop_names = merchant_df["shop_name"].dropna().astype(str).str.strip().tolist()
    merchant_names = merchant_df["merchant_name"].dropna().astype(str).str.strip().tolist()
    
    shop_mobile_map = dict(zip(merchant_df["shop_name"], merchant_df.get("merchant_mobile_number", pd.Series([""]*len(merchant_df)))))
    shop_state_map = dict(zip(merchant_df["shop_name"], merchant_df.get("shop_state", pd.Series([""]*len(merchant_df)))))
    merchant_mobile_map = dict(zip(merchant_df["merchant_name"], merchant_df.get("merchant_mobile_number", pd.Series([""]*len(merchant_df)))))
    merchant_state_map = dict(zip(merchant_df["merchant_name"], merchant_df.get("shop_state", pd.Series([""]*len(merchant_df)))))
    
    matched_merchants, buyer_mobiles, shop_states, merchant_scores, find_status = [], [], [], [], []
    
    for merchant in sale_order_df["merchant_name"]:
        match, score = exact_match_name(merchant, shop_names)
        if score == 100:
            matched_merchants.append(match)
            buyer_mobiles.append(shop_mobile_map.get(match, ""))
            shop_states.append(shop_state_map.get(match, ""))
            merchant_scores.append(score)
            find_status.append("shop_name")
            continue
        
        match2, score2 = exact_match_name(merchant, merchant_names)
        if score2 == 100:
            matched_merchants.append(match2)
            buyer_mobiles.append(merchant_mobile_map.get(match2, ""))
            shop_states.append(merchant_state_map.get(match2, ""))
            merchant_scores.append(score2)
            find_status.append("merchant_name")
        else:
            matched_merchants.append("")
            buyer_mobiles.append("")
            shop_states.append("")
            merchant_scores.append(0)
            find_status.append("not_found")
    
    sale_order_df["matched_shop_name"] = matched_merchants
    sale_order_df["buyer_mobile"] = buyer_mobiles
    sale_order_df["shop_state"] = shop_states
    sale_order_df["merchant_match_score"] = merchant_scores
    sale_order_df["merchant_find_status"] = find_status
    sale_order_df["merchant_error"] = [
        "Merchant not matched" if s == "not_found" else "" for s in sale_order_df["merchant_find_status"]
    ]
    
    # Combine errors
    sale_order_df["error_message"] = sale_order_df.apply(
        lambda r: ", ".join(filter(None, [r.get("product_error", ""), r.get("merchant_error", "")])),
        axis=1
    )
    
    # Separate negative quantities
    negative_qty_df = sale_order_df[sale_order_df["quantity"] < 0].copy()
    sale_order_df = sale_order_df[sale_order_df["quantity"] >= 0].copy()
    
    if not negative_qty_df.empty:
        negative_qty_df["abs_quantity"] = negative_qty_df["quantity"].abs()
        negative_qty_df["abs_net_sales"] = negative_qty_df["net_sales"].abs()
        negative_qty_df["price"] = negative_qty_df.apply(
            lambda row: row["abs_net_sales"] / row["abs_quantity"] 
            if row["abs_quantity"] != 0 and not pd.isna(row["abs_quantity"]) and not pd.isna(row["abs_net_sales"]) 
            else 0,
            axis=1
        )
        negative_qty_df["return_amount"] = negative_qty_df["price"]
        
        reason_col = negative_qty_df["low_price_reason"] if "low_price_reason" in negative_qty_df.columns else pd.Series(["low price"] * len(negative_qty_df))
        sales_return_error_message = negative_qty_df.apply(
            lambda r: ", ".join(filter(None, [r.get("product_error", ""), r.get("merchant_error", "")])),
            axis=1
        )
        product_name_col = negative_qty_df.get("matched_product_name", negative_qty_df.get("product_name", pd.Series([""] * len(negative_qty_df))))
        
        sales_return_df = pd.DataFrame({
            "order_id": negative_qty_df["order_id"],
            "product_id": negative_qty_df["product_id"],
            "batch_id": negative_qty_df["batch_id"],
            "product_name": product_name_col.astype(str),
            "price": negative_qty_df["price"],
            "return_qty": negative_qty_df["abs_quantity"],
            "return_amount": negative_qty_df["return_amount"],
            "reason": reason_col.astype(str),
            "error_message": sales_return_error_message.astype(str),
            "sales_return_date": negative_qty_df["order_date"],
            "note": "",
            "remark": ""
        })
    else:
        sales_return_df = pd.DataFrame()
    
    # Categorize orders
    def get_match_category(row):
        product_score = row.get("product_match_score", 0) or 0
        merchant_score = row.get("merchant_match_score", 0) or 0
        user_confirmed = row.get("user_confirmed_match", False)
        has_critical_error = bool(row.get("product_error") and "not found" in str(row.get("product_error", "")).lower()) or bool(row.get("merchant_error"))
        
        if ((product_score == 100 or user_confirmed) and merchant_score == 100 and not has_critical_error):
            return "valid"
        if (PARTIAL_MATCH_MIN_SCORE <= product_score < 100 and merchant_score == 100 and not has_critical_error and not user_confirmed):
            return "partial"
        return "error"
    
    sale_order_df["match_category"] = sale_order_df.apply(get_match_category, axis=1)
    
    rows_with_errors = sale_order_df[sale_order_df["error_message"].astype(str).str.strip() != ""]
    error_order_ids = rows_with_errors["order_id"].unique()
    
    rows_with_partial = sale_order_df[sale_order_df["match_category"] == "partial"]
    partial_order_ids = rows_with_partial["order_id"].unique()
    
    error_df = sale_order_df[sale_order_df["order_id"].isin(error_order_ids)].copy()
    orders_without_errors = sale_order_df[~sale_order_df["order_id"].isin(error_order_ids)].copy()
    partial_df = orders_without_errors[orders_without_errors["order_id"].isin(partial_order_ids)].copy()
    valid_df = orders_without_errors[~orders_without_errors["order_id"].isin(partial_order_ids)].copy()
    
    # Create error messages
    def mk_err_msg(r):
        parts = []
        if r.get("product_error"):
            parts.append(r["product_error"])
        if r.get("merchant_error"):
            parts.append(r["merchant_error"])
        product_score = r.get("product_match_score", 0) or 0
        merchant_score = r.get("merchant_match_score", 0) or 0
        user_confirmed = r.get("user_confirmed_match", False)
        if product_score < 100 and not user_confirmed:
            parts.append(f"Product match {int(product_score)}%")
        if merchant_score < 100:
            parts.append(f"Merchant match {int(merchant_score)}%")
        return ", ".join(parts) if parts else ""
    
    def mk_partial_msg(r):
        parts = []
        product_score = r.get("product_match_score", 0) or 0
        merchant_score = r.get("merchant_match_score", 0) or 0
        user_confirmed = r.get("user_confirmed_match", False)
        if user_confirmed:
            parts.append(f"User confirmed match ({int(product_score)}%)")
        elif product_score < 100:
            parts.append(f"Product match {int(product_score)}%")
        if merchant_score < 100:
            parts.append(f"Merchant match {int(merchant_score)}%")
        if r.get("product_error"):
            parts.append(r["product_error"])
        return ", ".join(parts) if parts else "Partial match"
    
    if not error_df.empty:
        error_df["error_message"] = error_df.apply(mk_err_msg, axis=1)
    
    if not partial_df.empty:
        partial_df["partial_match_reason"] = partial_df.apply(mk_partial_msg, axis=1)
    
    # Reorder columns
    top_cols = ["order_id","order_date","warehouse_name","product_id","batch_id",
                "buyer_mobile","buyer_branch_id","quantity","selling_price","due_date",
                "dms_invoice","low_price_reason","[OutputCGST Rate]","[OutputSGST Rate]"]
    
    if not valid_df.empty:
        rest_cols = [c for c in valid_df.columns if c not in top_cols]
        valid_df = valid_df[top_cols + rest_cols]
    
    if not partial_df.empty:
        rest_cols_partial = [c for c in partial_df.columns if c not in top_cols]
        partial_df = partial_df[top_cols + rest_cols_partial]
    
    # Save to Excel
    with pd.ExcelWriter(output_file, engine="openpyxl", mode="w") as writer:
        if not valid_df.empty:
            valid_df.to_excel(writer, sheet_name="Sale Order Demo", index=False)
        if not partial_df.empty:
            partial_df.to_excel(writer, sheet_name="Partially Matched", index=False)
        if not error_df.empty:
            error_df.to_excel(writer, sheet_name="Error Rows", index=False)
        if not sales_return_df.empty:
            sales_return_df.to_excel(writer, sheet_name="Sales Return Sheet", index=False)
        product_df.to_excel(writer, sheet_name="Product Details", index=False)
        merchant_df.to_excel(writer, sheet_name="merchant_data", index=False)
    
    # Apply color fills
    try:
        wb = load_workbook(output_file)
        if "Error Rows" in wb.sheetnames:
            ws = wb["Error Rows"]
            red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            for r in range(2, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = red
        if "Partially Matched" in wb.sheetnames:
            ws = wb["Partially Matched"]
            yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            for r in range(2, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = yellow
        wb.save(output_file)
    except Exception as e:
        if st:
            st.warning(f"Could not apply coloring to Excel file: {e}")
    
    return valid_df, partial_df, error_df, sales_return_df


