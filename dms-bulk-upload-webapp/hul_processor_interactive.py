"""
HUL Processor - Interactive version that asks questions during processing
Just like the original script, but with Streamlit buttons instead of input()
"""
import pandas as pd
import numpy as np
from fuzzywuzzy import process, fuzz
import os
from collections import Counter
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import streamlit as st

# Utility functions (same as original)
def normalize_name(name):
    if pd.isna(name):
        return ""
    cleaned = " ".join(str(name).strip().split())
    return cleaned

def fuzzy_match_name(name, choices, min_score=0):
    normalized_input = normalize_name(name)
    if not normalized_input:
        return "", 0
    normalized_choices = {normalize_name(c): c for c in choices if normalize_name(c)}
    normalized_choice_list = list(normalized_choices.keys())
    result = process.extractOne(normalized_input, normalized_choice_list, scorer=fuzz.token_sort_ratio)
    if result:
        match_normalized, score = result
        match = normalized_choices.get(match_normalized, match_normalized)
        if score >= min_score:
            return match, score
        else:
            return "", 0
    return "", 0

def exact_match_name(name, choices):
    normalized_input = normalize_name(name).lower()
    if not normalized_input:
        return "", 0
    for choice in choices:
        normalized_choice = normalize_name(choice).lower()
        if normalized_input == normalized_choice:
            return choice, 100
    return "", 0

def safe_read_excel(file_path, **kwargs):
    ext = os.path.splitext(file_path)[-1].lower()
    if ext == ".xls":
        return pd.read_excel(file_path, engine="xlrd", **kwargs)
    elif ext == ".xlsx":
        return pd.read_excel(file_path, engine="openpyxl", **kwargs)
    else:
        raise ValueError(f"Unsupported file type: {ext}")

# Import shared utility for interactive questions
try:
    from interactive_utils import ask_user_question_streamlit
except ImportError:
    # Fallback if interactive_utils not available
    def ask_user_question_streamlit(question_type, question_data, processing_cache):
        cache_key = question_data['cache_key']
        if question_type == 'partial_match':
            if cache_key in processing_cache['partial_matches']:
                return processing_cache['partial_matches'][cache_key]
        elif question_type == 'variant':
            if cache_key in processing_cache['variants']:
                return processing_cache['variants'][cache_key]
        elif question_type == 'related':
            if cache_key in processing_cache['related']:
                return processing_cache['related'][cache_key]
        if st:
            st.session_state.pending_question = {
                'type': question_type,
                'cache_key': cache_key,
                **question_data
            }
            return None
        return False

def process_hul_sales_interactive(input_file, reference_file, output_file, 
                                 column_mapping=None, selected_state=None,
                                 processing_cache=None, pending_question_callback=None):
    """
    Process HUL sales orders - Interactive version that asks questions during processing.
    Works just like the original script - asks questions as they come up.
    """
    PARTIAL_MATCH_MIN_SCORE = 70
    
    # Initialize cache if not provided
    if processing_cache is None:
        processing_cache = {
            'partial_matches': {},
            'variants': {},
            'related': {}
        }
    
    # Load sales file
    df = safe_read_excel(input_file)
    
    # Column mapping (auto-detect)
    if column_mapping is None:
        expected = {
            "order_id": "Bill Number",
            "dms_invoice": "Bill Number",
            "order_date": "Bill Date",
            "product_name": "Product Description",
            "merchant_name": "Party",
            "quantity": "Units",
            "selling_price": "Net Sales",
            "low_price_reason": "Low Price Reason",
            "buyer_branch_id": "Branch ID",
            "warehouse_name": "Warehouse Name"
        }
        
        mapping = {}
        optional_columns = ["low_price_reason", "buyer_branch_id", "warehouse_name"]
        
        for key, default in expected.items():
            if key in optional_columns:
                if default in df.columns:
                    mapping[key] = default
                else:
                    mapping[key] = None
            else:
                if default in df.columns:
                    mapping[key] = default
                else:
                    for col in df.columns:
                        if key.lower().replace("_", " ") in str(col).lower():
                            mapping[key] = col
                            break
                    if key not in mapping:
                        mapping[key] = df.columns[0] if len(df.columns) > 0 else None
    else:
        mapping = column_mapping
    
    # Detect due_date and Total Tax % columns
    due_date_col = None
    total_tax_col = None
    for col in df.columns:
        if "due" in str(col).strip().lower():
            due_date_col = col
            break
    
    # Detect Total Tax % column (with fallback)
    for col in df.columns:
        if "total tax" in str(col).strip().lower() and "%" in str(col):
            total_tax_col = col
            break
    if not total_tax_col:
        # Try alternative names
        for col in df.columns:
            if "total tax" in str(col).strip().lower() or "tax %" in str(col).strip().lower():
                total_tax_col = col
                break
    
    # Build sale_order_df
    sale_order_df = pd.DataFrame({
        "order_id": df[mapping["order_id"]],
        "dms_invoice": df[mapping["dms_invoice"]],
        "order_date": pd.to_datetime(df[mapping["order_date"]], errors="coerce").dt.strftime("%d/%m/%Y"),
        "product_name": df[mapping["product_name"]].astype(str).str.strip(),
        "merchant_name": df[mapping["merchant_name"]].astype(str).str.strip(),
        "quantity": pd.to_numeric(df[mapping["quantity"]], errors="coerce"),
        "net_sales": pd.to_numeric(df[mapping["selling_price"]], errors="coerce"),
        "low_price_reason": df[mapping.get("low_price_reason")].astype(str) 
                        if mapping.get("low_price_reason") and mapping.get("low_price_reason") in df.columns 
                        else pd.Series(["low price"] * len(df)),
        "buyer_branch_id": df[mapping.get("buyer_branch_id")].astype(str) 
                        if mapping.get("buyer_branch_id") and mapping.get("buyer_branch_id") in df.columns 
                        else pd.Series([""] * len(df)),
        "warehouse_name": df[mapping.get("warehouse_name")].astype(str) 
                        if mapping.get("warehouse_name") and mapping.get("warehouse_name") in df.columns 
                        else pd.Series([""] * len(df))
    })
    
    if due_date_col:
        sale_order_df["due_date"] = pd.to_datetime(df[due_date_col], errors="coerce").dt.strftime("%d/%m/%Y")
    else:
        sale_order_df["due_date"] = ""
    
    if total_tax_col:
        sale_order_df["Total Tax %"] = pd.to_numeric(df[total_tax_col], errors="coerce")
    else:
        sale_order_df["Total Tax %"] = ""
    
    # Drop incomplete rows
    before = len(sale_order_df)
    sale_order_df = sale_order_df.dropna(subset=["order_id", "order_date", "product_name", "merchant_name"])
    if st:
        st.info(f"Dropped {before - len(sale_order_df)} incomplete rows")
    
    # Calculate selling_price
    sale_order_df["selling_price"] = sale_order_df.apply(
        lambda row: row["net_sales"] / row["quantity"] if row["quantity"] != 0 and not pd.isna(row["quantity"]) and not pd.isna(row["net_sales"]) else 0,
        axis=1
    )
    
    # Load reference sheets
    product_df = safe_read_excel(reference_file, sheet_name="Product Details", dtype={"batch_id": str})
    merchant_df = safe_read_excel(reference_file, sheet_name="merchant_data")
    
    # Clean batch_id
    if "batch_id" in product_df.columns:
        product_df["batch_id"] = product_df["batch_id"].apply(lambda x: str(x).split(".")[0] if pd.notna(x) else "")
    
    # State selection
    unique_states = merchant_df["shop_state"].dropna().unique().tolist()
    if not unique_states:
        raise RuntimeError("No states found in merchant_data.shop_state")
    
    if selected_state is None:
        selected_state = unique_states[0]
    
    merchant_df = merchant_df[merchant_df["shop_state"] == selected_state].copy()
    
    # Check if product matching is already complete
    product_matching_complete = False
    matched_products, product_scores, user_confirmed_flags = [], [], []  # Initialize
    
    if st and 'processing_phase' in st.session_state:
        if st.session_state.processing_phase == 'batch_allocation':
            product_matching_complete = True
            # Restore matched products from session state
            if 'matched_products_data' in st.session_state:
                matched_products = st.session_state.matched_products_data.get('matched_products', []).copy()
                product_scores = st.session_state.matched_products_data.get('product_scores', []).copy()
                user_confirmed_flags = st.session_state.matched_products_data.get('user_confirmed_flags', []).copy()
            else:
                # Phase says complete but no data - need to redo matching
                product_matching_complete = False
                matched_products, product_scores, user_confirmed_flags = [], [], []
    
    # Product matching with interactive confirmation (like original script)
    if not product_matching_complete:
        product_names = product_df["product_name"].dropna().astype(str).str.strip().tolist()
        
        # Initialize or restore matched products from session state
        if st and 'matched_products_data' in st.session_state:
            matched_products = st.session_state.matched_products_data.get('matched_products', []).copy()
            product_scores = st.session_state.matched_products_data.get('product_scores', []).copy()
            user_confirmed_flags = st.session_state.matched_products_data.get('user_confirmed_flags', []).copy()
        else:
            matched_products, product_scores, user_confirmed_flags = [], [], []
        
        progress_bar = st.progress(0) if st else None
        status_text = st.empty() if st else None
        
        total_products = len(sale_order_df)
        start_idx = len(matched_products)  # Resume from where we left off
        
        # Debug: show where we're resuming from
        if st and start_idx > 0:
            st.info(f"🔄 Resuming product matching from {start_idx + 1}/{total_products}")
        
        for idx, prod in enumerate(sale_order_df["product_name"]):
            # Skip already processed products
            if idx < start_idx:
                continue
            
            match, score = fuzzy_match_name(prod, product_names, min_score=0)
            user_confirmed = False
            
            # Check if this is a partial match (70-99%) that needs user confirmation
            if match and 70 <= score < 100:
                cache_key = f"{prod}|{match}"
                
                # Use cache only - no questions during processing
                if cache_key in processing_cache['partial_matches']:
                    user_confirmed = processing_cache['partial_matches'][cache_key]
                else:
                    # Not in cache - default to False (reject) since all questions should be answered upfront
                    user_confirmed = False
                    processing_cache['partial_matches'][cache_key] = False
            elif score >= 100:
                user_confirmed = True
            
            matched_products.append(match)
            product_scores.append(score)
            user_confirmed_flags.append(user_confirmed)
            
            # Update session state with progress
            if st:
                st.session_state.matched_products_data = {
                    'matched_products': matched_products,
                    'product_scores': product_scores,
                    'user_confirmed_flags': user_confirmed_flags
                }
            
            if progress_bar and status_text:
                progress = (idx + 1) / total_products
                progress_bar.progress(progress)
                status_text.text(f"Matching products: {idx + 1}/{total_products}")
        
        # Mark product matching as complete and save to session state
        if st:
            st.session_state.processing_phase = 'batch_allocation'  # Move to next phase
            st.session_state.matched_products_data = {
                'matched_products': matched_products.copy(),
                'product_scores': product_scores.copy(),
                'user_confirmed_flags': user_confirmed_flags.copy()
            }
        
        if progress_bar:
            progress_bar.empty()
        if status_text:
            status_text.empty()
    else:
        # Product matching already complete - restore from session state
        if st:
            st.info("✅ Product matching already complete. Continuing with batch allocation...")
    
    # Set matched products in dataframe
    sale_order_df["matched_product_name"] = matched_products
    sale_order_df["product_match_score"] = product_scores
    sale_order_df["user_confirmed_match"] = user_confirmed_flags
    
    # Build batch inventory and product variants
    batch_inventory = {}
    product_id_groups = {}
    
    for _, r in product_df.iterrows():
        pname = str(r["product_name"]).strip()
        pid = str(r.get("product_id", ""))
        
        batch_inventory.setdefault(pname, []).append({
            "batch_id": str(r.get("batch_id", "")),
            "available_stock": float(r.get("available_stock", np.inf)) if pd.notna(r.get("available_stock", np.nan)) else np.inf,
            "product_id": pid,
        })
        
        if pid:
            product_id_groups.setdefault(pid, []).append(pname)
    
    # Build product variants map
    product_variants = {}
    for pid, names in product_id_groups.items():
        if len(names) > 1:
            unique_names = list(set(names))
            if len(unique_names) > 1:
                unique_names.sort()
                main_product = unique_names[0]
                variants = unique_names[1:]
                product_variants[main_product] = variants
    
    for pname in batch_inventory:
        batch_inventory[pname] = sorted(batch_inventory[pname], key=lambda x: -x["available_stock"])
    
    # Batch allocation with interactive confirmations (like original script)
    sale_order_df = sale_order_df.sort_values(by="order_date").reset_index(drop=True)
    allocated_rows = []
    
    progress_bar2 = st.progress(0) if st else None
    status_text2 = st.empty() if st else None
    
    for idx, row in enumerate(sale_order_df.itertuples()):
        pname = row.matched_product_name
        qty = row.quantity if not pd.isna(row.quantity) else 0
        
        base_row = {
            "order_id": row.order_id,
            "dms_invoice": row.dms_invoice,
            "order_date": row.order_date,
            "product_name": row.product_name,
            "merchant_name": row.merchant_name,
            "net_sales": row.net_sales,
            "low_price_reason": row.low_price_reason,
            "buyer_branch_id": row.buyer_branch_id,
            "warehouse_name": row.warehouse_name,
            "due_date": row.due_date,
            "matched_product_name": row.matched_product_name,
            "product_match_score": row.product_match_score,
            "user_confirmed_match": row.user_confirmed_match,
            "Total Tax %": sale_order_df.loc[row.Index, "Total Tax %"] if "Total Tax %" in sale_order_df.columns else "",
        }
        
        if not pname:
            if pd.notna(row.product_match_score) and row.product_match_score > 0:
                err = f"Product match score too low ({int(row.product_match_score)}%) - may be similar but different product"
            else:
                err = "Product not found in reference"
            
            allocated_rows.append({
                **base_row,
                "quantity": qty,
                "selling_price": row.net_sales / qty if qty != 0 else 0,
                "batch_id": "",
                "product_id": "",
                "product_error": err
            })
        elif pname not in batch_inventory:
            err = "Product not found in reference"
            allocated_rows.append({
                **base_row,
                "quantity": qty,
                "selling_price": row.net_sales / qty if qty != 0 else 0,
                "batch_id": "",
                "product_id": "",
                "product_error": err
            })
        elif qty < 0:
            # Returns
            batch = batch_inventory[pname][0]
            batch["available_stock"] += abs(qty)
            allocated_rows.append({
                **base_row,
                "quantity": qty,
                "selling_price": row.net_sales / qty if qty != 0 else 0,
                "batch_id": batch["batch_id"],
                "product_id": batch["product_id"],
                "product_error": ""
            })
        elif qty > 0:
            # Multi-batch allocation with interactive confirmations
            products_to_use = [pname]
            current_stock = sum(b["available_stock"] for b in batch_inventory[pname])
            
            # STEP 2: Check for product variants if stock is insufficient
            if current_stock < qty:
                variants_to_check = []
                if pname in product_variants:
                    variants_to_check = product_variants[pname]
                else:
                    for main_prod, variants in product_variants.items():
                        if pname in variants:
                            variants_to_check = [main_prod] + [v for v in variants if v != pname]
                            break
                
                # Filter variants with available stock
                variants_with_stock = []
                for variant in variants_to_check:
                    if variant in batch_inventory:
                        variant_stock = sum(b["available_stock"] for b in batch_inventory[variant])
                        if variant_stock > 0:
                            variants_with_stock.append((variant, variant_stock))
                
                # Use cache only - no questions during processing
                for variant, variant_stock in variants_with_stock:
                    cache_key = f"{pname}|{variant}"
                    
                    # Check cache
                    if cache_key in processing_cache['variants']:
                        if processing_cache['variants'][cache_key]:
                            products_to_use.append(variant)
                    else:
                        # Not in cache - default to False (don't use) since all questions should be answered upfront
                        processing_cache['variants'][cache_key] = False
            
            # STEP 3: Check for related products if still insufficient
            total_stock = sum(sum(b["available_stock"] for b in batch_inventory[p]) for p in products_to_use)
            
            if total_stock < qty:
                pname_lower = pname.lower().strip()
                related_products = []
                
                for prod_key in batch_inventory.keys():
                    if prod_key in products_to_use:
                        continue
                    prod_lower = prod_key.lower().strip()
                    similarity = fuzz.ratio(pname_lower, prod_lower)
                    
                    is_potentially_related = False
                    if len(pname_lower) >= 10 and len(prod_lower) >= 10:
                        if (pname_lower in prod_lower) or (prod_lower in pname_lower):
                            is_potentially_related = True
                    if not is_potentially_related and similarity >= 80:
                        is_potentially_related = True
                    
                    if is_potentially_related:
                        rp_stock = sum(b["available_stock"] for b in batch_inventory[prod_key])
                        if rp_stock > 0:
                            related_products.append((prod_key, rp_stock))
                
                # Use cache only - no questions during processing
                for rp, rp_stock in related_products:
                    cache_key = f"{pname}|{rp}"
                    
                    # Check cache
                    if cache_key in processing_cache['related']:
                        if processing_cache['related'][cache_key]:
                            products_to_use.append(rp)
                    else:
                        # Not in cache - default to False (different product) since all questions should be answered upfront
                        processing_cache['related'][cache_key] = False
            
            # STEP 4: Allocate from all selected products using MULTIPLE BATCHES
            remaining_qty = qty
            batch_allocations = []
            
            # Collect all batches from selected products
            all_batches = []
            for prod in products_to_use:
                if prod in batch_inventory:
                    for batch in batch_inventory[prod]:
                        all_batches.append(batch)
            
            # Sort by available stock descending
            all_batches = sorted(all_batches, key=lambda x: -x["available_stock"])
            
            # Allocate from multiple batches until order is fulfilled
            for batch in all_batches:
                if remaining_qty <= 0:
                    break
                
                if batch["available_stock"] > 0:
                    qty_from_batch = min(batch["available_stock"], remaining_qty)
                    batch["available_stock"] -= qty_from_batch
                    remaining_qty -= qty_from_batch
                    batch_allocations.append((batch["batch_id"], batch["product_id"], qty_from_batch))
            
            if remaining_qty <= 0:
                # Success! Create one row per batch allocation
                for batch_id, product_id, allocated_qty in batch_allocations:
                    portion_selling_price = (row.net_sales * allocated_qty / qty) if qty != 0 else 0
                    allocated_rows.append({
                        **base_row,
                        "quantity": allocated_qty,
                        "selling_price": portion_selling_price / allocated_qty if allocated_qty != 0 else 0,
                        "batch_id": batch_id,
                        "product_id": product_id,
                        "product_error": ""
                    })
                
                if len(batch_allocations) > 1:
                    if st:
                        st.info(f"✅ Order {row.order_id} split into {len(batch_allocations)} rows (batches: {', '.join([b[0] for b in batch_allocations])}) | Product: '{pname}' | Total: {int(qty)} units")
            else:
                fulfilled_qty = qty - remaining_qty
                err = f"Insufficient stock: need {int(qty)}, only {int(fulfilled_qty)} available across all batches (matched to: '{pname}')"
                allocated_rows.append({
                    **base_row,
                    "quantity": qty,
                    "selling_price": row.net_sales / qty if qty != 0 else 0,
                    "batch_id": "",
                    "product_id": "",
                    "product_error": err
                })
        else:
            allocated_rows.append({
                **base_row,
                "quantity": qty,
                "selling_price": 0,
                "batch_id": "",
                "product_id": "",
                "product_error": "Zero quantity"
            })
        
        if progress_bar2 and status_text2:
            progress = (idx + 1) / len(sale_order_df)
            progress_bar2.progress(progress)
            status_text2.text(f"Allocating batches: {idx + 1}/{len(sale_order_df)}")
    
    if progress_bar2:
        progress_bar2.empty()
    if status_text2:
        status_text2.empty()
    
    sale_order_df = pd.DataFrame(allocated_rows)
    
    # Merchant matching (same as before)
    shop_names = merchant_df["shop_name"].dropna().astype(str).str.strip().tolist()
    merchant_names = merchant_df["merchant_name"].dropna().astype(str).str.strip().tolist()
    
    shop_mobile_map = dict(zip(merchant_df["shop_name"], merchant_df.get("merchant_mobile_number", pd.Series([""]*len(merchant_df)))))
    shop_state_map = dict(zip(merchant_df["shop_name"], merchant_df.get("shop_state", pd.Series([""]*len(merchant_df)))))
    merchant_mobile_map = dict(zip(merchant_df["merchant_name"], merchant_df.get("merchant_mobile_number", pd.Series([""]*len(merchant_df)))))
    merchant_state_map = dict(zip(merchant_df["merchant_name"], merchant_df.get("shop_state", pd.Series([""]*len(merchant_df)))))
    
    matched_merchants, buyer_mobiles, shop_states, merchant_scores, find_status = [], [], [], [], []
    
    for merchant in sale_order_df["merchant_name"]:
        match, score = exact_match_name(merchant, shop_names)
        if score == 100:
            matched_merchants.append(match)
            buyer_mobiles.append(shop_mobile_map.get(match, ""))
            shop_states.append(shop_state_map.get(match, ""))
            merchant_scores.append(score)
            find_status.append("shop_name")
            continue
        
        match2, score2 = exact_match_name(merchant, merchant_names)
        if score2 == 100:
            matched_merchants.append(match2)
            buyer_mobiles.append(merchant_mobile_map.get(match2, ""))
            shop_states.append(merchant_state_map.get(match2, ""))
            merchant_scores.append(score2)
            find_status.append("merchant_name")
        else:
            matched_merchants.append("")
            buyer_mobiles.append("")
            shop_states.append("")
            merchant_scores.append(0)
            find_status.append("not_found")
    
    sale_order_df["matched_shop_name"] = matched_merchants
    sale_order_df["buyer_mobile"] = buyer_mobiles
    sale_order_df["shop_state"] = shop_states
    sale_order_df["merchant_match_score"] = merchant_scores
    sale_order_df["merchant_find_status"] = find_status
    sale_order_df["merchant_error"] = [
        "Merchant not matched" if s == "not_found" else "" for s in sale_order_df["merchant_find_status"]
    ]
    
    # Combine errors
    sale_order_df["error_message"] = sale_order_df.apply(
        lambda r: ", ".join(filter(None, [r.get("product_error", ""), r.get("merchant_error", "")])),
        axis=1
    )
    
    # Separate negative quantities
    negative_qty_df = sale_order_df[sale_order_df["quantity"] < 0].copy()
    sale_order_df = sale_order_df[sale_order_df["quantity"] >= 0].copy()
    
    if not negative_qty_df.empty:
        negative_qty_df["abs_quantity"] = negative_qty_df["quantity"].abs()
        negative_qty_df["abs_net_sales"] = negative_qty_df["net_sales"].abs()
        negative_qty_df["price"] = negative_qty_df.apply(
            lambda row: row["abs_net_sales"] / row["abs_quantity"] 
            if row["abs_quantity"] != 0 and not pd.isna(row["abs_quantity"]) and not pd.isna(row["abs_net_sales"]) 
            else 0,
            axis=1
        )
        negative_qty_df["return_amount"] = negative_qty_df["price"]
        
        reason_col = negative_qty_df["low_price_reason"] if "low_price_reason" in negative_qty_df.columns else pd.Series(["low price"] * len(negative_qty_df))
        sales_return_error_message = negative_qty_df.apply(
            lambda r: ", ".join(filter(None, [r.get("product_error", ""), r.get("merchant_error", "")])),
            axis=1
        )
        product_name_col = negative_qty_df.get("matched_product_name", negative_qty_df.get("product_name", pd.Series([""] * len(negative_qty_df))))
        
        sales_return_df = pd.DataFrame({
            "order_id": negative_qty_df["order_id"],
            "product_id": negative_qty_df["product_id"],
            "batch_id": negative_qty_df["batch_id"],
            "product_name": product_name_col.astype(str),
            "price": negative_qty_df["price"],
            "return_qty": negative_qty_df["abs_quantity"],
            "return_amount": negative_qty_df["return_amount"],
            "reason": reason_col.astype(str),
            "error_message": sales_return_error_message.astype(str),
            "sales_return_date": negative_qty_df["order_date"],
            "note": "",
            "remark": ""
        })
    else:
        sales_return_df = pd.DataFrame()
    
    # Categorize orders
    # STRICT RULE: Only 100% product AND 100% merchant matches go to "Sale Order Demo"
    def get_match_category(row):
        product_score = row.get("product_match_score", 0) or 0
        merchant_score = row.get("merchant_match_score", 0) or 0
        user_confirmed = row.get("user_confirmed_match", False)
        has_critical_error = bool(row.get("product_error") and "not found" in str(row.get("product_error", "")).lower()) or bool(row.get("merchant_error"))
        
        # Check for incomplete orders (missing product_id or batch_id)
        product_id = row.get("product_id", "")
        batch_id = row.get("batch_id", "")
        is_incomplete = (not product_id or product_id == "" or pd.isna(product_id)) or (not batch_id or batch_id == "" or pd.isna(batch_id))
        
        # Valid: ONLY 100% product AND 100% merchant, no errors, and complete
        if product_score == 100 and merchant_score == 100 and not has_critical_error and not is_incomplete:
            return "valid"
        
        # Partial: Product 70-99% AND merchant 100%, no errors, complete, and user confirmed
        if user_confirmed and PARTIAL_MATCH_MIN_SCORE <= product_score < 100 and merchant_score == 100 and not has_critical_error and not is_incomplete:
            return "valid"  # User confirmed partial goes to valid
        
        # Partial: Product 70-99% AND merchant 100%, no errors, complete, NOT user confirmed
        if PARTIAL_MATCH_MIN_SCORE <= product_score < 100 and merchant_score == 100 and not has_critical_error and not is_incomplete and not user_confirmed:
            return "partial"
        
        # Error: Everything else (product <70%, merchant <100%, incomplete, or has errors)
        return "error"
    
    sale_order_df["match_category"] = sale_order_df.apply(get_match_category, axis=1)
    
    # Find order_ids with any errors: error_message OR match_category == "error"
    # This ensures orders with low match scores (<70%) or incomplete orders go to error sheet
    rows_with_errors = sale_order_df[
        (sale_order_df["error_message"].astype(str).str.strip() != "") | 
        (sale_order_df["match_category"] == "error")
    ]
    error_order_ids = rows_with_errors["order_id"].unique()
    
    # Find order_ids with any partial matches (move entire order to partial sheet)
    rows_with_partial = sale_order_df[sale_order_df["match_category"] == "partial"]
    partial_order_ids = rows_with_partial["order_id"].unique()
    
    # Create three dataframes - ENTIRE orders move together
    # If any row in an order has an error, move ALL rows of that order to error sheet
    error_df = sale_order_df[sale_order_df["order_id"].isin(error_order_ids)].copy()
    
    # Exclude error orders first, then check for partial orders
    orders_without_errors = sale_order_df[~sale_order_df["order_id"].isin(error_order_ids)].copy()
    
    # If ANY product in an order is "partial", move ENTIRE order to Partially Matched
    partial_df = orders_without_errors[orders_without_errors["order_id"].isin(partial_order_ids)].copy()
    
    # Remaining orders (all products are valid) go to Valid sheet
    valid_df = orders_without_errors[~orders_without_errors["order_id"].isin(partial_order_ids)].copy()
    
    # Create error messages
    def mk_err_msg(r):
        parts = []
        if r.get("product_error"):
            parts.append(r["product_error"])
        if r.get("merchant_error"):
            parts.append(r["merchant_error"])
        product_score = r.get("product_match_score", 0) or 0
        merchant_score = r.get("merchant_match_score", 0) or 0
        user_confirmed = r.get("user_confirmed_match", False)
        if product_score < 100 and not user_confirmed:
            parts.append(f"Product match {int(product_score)}%")
        if merchant_score < 100:
            parts.append(f"Merchant match {int(merchant_score)}%")
        return ", ".join(parts) if parts else ""
    
    def mk_partial_msg(r):
        parts = []
        product_score = r.get("product_match_score", 0) or 0
        merchant_score = r.get("merchant_match_score", 0) or 0
        user_confirmed = r.get("user_confirmed_match", False)
        if user_confirmed:
            parts.append(f"User confirmed match ({int(product_score)}%)")
        elif product_score < 100:
            parts.append(f"Product match {int(product_score)}%")
        if merchant_score < 100:
            parts.append(f"Merchant match {int(merchant_score)}%")
        if r.get("product_error"):
            parts.append(r["product_error"])
        return ", ".join(parts) if parts else "Partial match"
    
    if not error_df.empty:
        error_df["error_message"] = error_df.apply(mk_err_msg, axis=1)
    
    if not partial_df.empty:
        partial_df["partial_match_reason"] = partial_df.apply(mk_partial_msg, axis=1)
    
    # Reorder columns
    top_cols = ["order_id","order_date","warehouse_name","product_id","batch_id",
                "buyer_mobile","buyer_branch_id","quantity","selling_price","due_date",
                "dms_invoice","low_price_reason","Total Tax %"]
    
    if not valid_df.empty:
        rest_cols = [c for c in valid_df.columns if c not in top_cols]
        valid_df = valid_df[top_cols + rest_cols]
    
    if not partial_df.empty:
        rest_cols_partial = [c for c in partial_df.columns if c not in top_cols]
        partial_df = partial_df[top_cols + rest_cols_partial]
    
    # Split valid orders into multiple sheets if > 200 orders
    from sheet_splitter import split_orders_into_sheets
    valid_sheets, sheet_info = split_orders_into_sheets(valid_df)
    
    # Save to Excel
    with pd.ExcelWriter(output_file, engine="openpyxl", mode="w") as writer:
        # Write valid orders (split into multiple sheets if needed)
        for sheet_name, sheet_df in valid_sheets:
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        if not partial_df.empty:
            partial_df.to_excel(writer, sheet_name="Partially Matched", index=False)
        if not error_df.empty:
            error_df.to_excel(writer, sheet_name="Error Rows", index=False)
        if not sales_return_df.empty:
            sales_return_df.to_excel(writer, sheet_name="Sales Return Sheet", index=False)
        product_df.to_excel(writer, sheet_name="Product Details", index=False)
        merchant_df.to_excel(writer, sheet_name="merchant_data", index=False)
    
    # Apply color fills
    try:
        wb = load_workbook(output_file)
        if "Error Rows" in wb.sheetnames:
            ws = wb["Error Rows"]
            red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            for r in range(2, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = red
        if "Partially Matched" in wb.sheetnames:
            ws = wb["Partially Matched"]
            yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            for r in range(2, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = yellow
        wb.save(output_file)
    except Exception as e:
        if st:
            st.warning(f"Could not apply coloring to Excel file: {e}")
    
    # Store sheet info for summary display
    st.session_state.output_sheet_info = sheet_info if 'output_sheet_info' not in st.session_state else sheet_info
    
    # Summary statistics
    # Display summary like raw script
    if st:
        st.success("✅ Processing complete!")
        st.markdown("---")
        st.markdown("## 📊 SUMMARY")
        
        from collections import Counter
        
        valid_count = len(valid_df['order_id'].unique()) if not valid_df.empty else 0
        partial_count = len(partial_df['order_id'].unique()) if not partial_df.empty else 0
        error_count = len(error_df['order_id'].unique()) if not error_df.empty else 0
        sales_return_count = len(sales_return_df) if not sales_return_df.empty else 0
        
        st.markdown(f"  ✅ **Valid orders (100% match):** {valid_count}")
        st.markdown(f"  ⚠️  **Partially matched orders (70-99%):** {partial_count}")
        st.markdown(f"  ❌ **Orders with errors (<70%):** {error_count}")
        if sales_return_count > 0:
            st.markdown(f"  🔄 **Sales returns (count):** {sales_return_count}")
        
        # Show file breakdown if orders were split
        if sheet_info and len(sheet_info) > 1:
            st.markdown("")
            st.markdown("### 📄 Output File Breakdown")
            st.markdown(f"  **Total valid orders:** {valid_count} (split into {len(sheet_info)} sheets)")
            for info in sheet_info:
                st.markdown(f"  - **{info['sheet_name']}:** {info['order_count']} orders ({info['row_count']} rows)")
            st.info("💡 Each order is kept complete in one sheet - no orders are split across sheets.")
        
        # Multi-batch allocation stats
        if not valid_df.empty:
            order_counts = valid_df["order_id"].value_counts()
            split_orders = order_counts[order_counts > 1]
            if not split_orders.empty:
                total_splits = split_orders.sum()
                st.markdown("")
                st.markdown("### 📦 MULTI-BATCH ALLOCATION")
                st.markdown(f"  {len(split_orders)} orders split into multiple rows (total {int(total_splits)} rows)")
                st.markdown("  This helped fulfill orders when single batches had insufficient stock")
                st.markdown("  Example: Order requiring 30 units might use Batch1(20) + Batch2(10) = 2 rows")
        
        # Partial match summary
        if not partial_df.empty:
            st.markdown("")
            st.markdown("### ⚠️  Partial match summary (top reasons):")
            for reason, cnt in Counter(partial_df["partial_match_reason"]).most_common(5):
                st.markdown(f"  - {reason}: {cnt}")
        
        # Error summary
        if not error_df.empty:
            st.markdown("")
            st.markdown("### ❌ Error summary (top reasons):")
            for reason, cnt in Counter(error_df["error_message"]).most_common(5):
                st.markdown(f"  - {reason}: {cnt}")
    
    return True  # Processing complete

